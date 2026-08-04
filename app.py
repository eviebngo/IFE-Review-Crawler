import csv
import io
import json
import os
import re
import socket
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path
from flask import Flask, render_template, request, jsonify, Response

try:
    import anthropic
    _ANTHROPIC_OK = True
except ImportError:
    _ANTHROPIC_OK = False

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent / ".env")
except ImportError:
    pass
from ife_crawler import (IFECrawler, AIRLINE_IFE_LOOKUP, IFE_FEATURE_KEYWORDS,
                         AIRLINE_KEYWORDS, AIRCRAFT_KEYWORDS)
from ife_data_manager import IFEDataManager

app = Flask(__name__)
app.config['JSON_SORT_KEYS'] = False

data_manager = IFEDataManager()
PER_PAGE = 50

# ── Background auto-discovery ─────────────────────────────────────────────────

# Interval between discovery runs (seconds). First run is immediate.
CRAWL_INTERVAL = 86400  # re-crawl once per day

_crawl_status = {
    "running":      False,
    "last_run":     None,
    "last_added":   0,
    "next_run":     None,
    "error":        None,
}


def _run_auto_discovery():
    """Background thread: search for IFE reviews by keyword and cache new ones."""
    global _crawl_status
    while True:
        _crawl_status["running"] = True
        _crawl_status["error"] = None
        try:
            data_manager.reload_from_disk()
            existing = {r["url"] for r in data_manager.data.get("reviews", [])}

            crawler = IFECrawler(verify_ssl=False, api_key=os.environ.get("YOUTUBE_API_KEY", ""))
            new_results = crawler.auto_discover(existing_urls=existing, max_results=500, days_lookback=7)

            if new_results:
                data_manager.data["reviews"].extend(new_results)
                data_manager.save_cache()

            _crawl_status["last_added"] = len(new_results)
            _crawl_status["last_run"] = datetime.now().isoformat()
        except Exception as e:
            _crawl_status["error"] = str(e)
        finally:
            _crawl_status["running"] = False
            _crawl_status["next_run"] = datetime.fromtimestamp(
                time.time() + CRAWL_INTERVAL
            ).isoformat()

        time.sleep(CRAWL_INTERVAL)


def start_background_crawler():
    t = threading.Thread(target=_run_auto_discovery, daemon=True)
    t.start()


def _purge_tier3_articles():
    """One-time startup: remove Tier 3 articles from the JSON cache file."""
    data_manager.reload_from_disk()
    before = len(data_manager.load_cache().get("reviews", []))
    kept   = len(data_manager.data.get("reviews", []))
    if kept < before:
        data_manager.save_cache()


def _initial_seed():
    """If the database is small, do a one-time broad crawl before the daily loop starts."""
    data_manager.reload_from_disk()
    if len(data_manager.data.get("reviews", [])) < 50:
        try:
            existing = {r["url"] for r in data_manager.data.get("reviews", [])}
            crawler  = IFECrawler(verify_ssl=False, api_key=os.environ.get("YOUTUBE_API_KEY", ""))
            results  = crawler.auto_discover(existing_urls=existing, max_results=500, days_lookback=365)
            if results:
                data_manager.reload_from_disk()
                data_manager.data["reviews"].extend(results)
                data_manager.save_cache()
        except Exception as e:
            print(f"[seed] error: {e}")


# Start on import (Flask dev-mode forks; only start once via the werkzeug check)
if not os.environ.get("WERKZEUG_RUN_MAIN") == "false":
    _purge_tier3_articles()
    threading.Thread(target=_initial_seed, daemon=True).start()
    start_background_crawler()


# ── Routes ────────────────────────────────────────────────────────────────────

def _request_is_local():
    """True only when the request originates from the machine running the server
    (loopback, or the host's own LAN address when browsing via the share link)."""
    ra = request.remote_addr or ""
    if ra in ("127.0.0.1", "::1"):
        return True
    try:
        infos = socket.getaddrinfo(socket.gethostname(), None)
        return ra in {i[4][0] for i in infos}
    except Exception:
        return False


@app.route("/")
def index():
    return render_template("index.html", can_edit=_request_is_local())


@app.route("/api/ife-reviews")
def get_ife_reviews():
    """Paginated full review list."""
    try:
        page = int(request.args.get("page", 1))
        data_manager.reload_from_disk()
        reviews = sorted(
            data_manager.data.get("reviews", []),
            key=data_manager._published_ts,
            reverse=True
        )
        paged = data_manager.paginate(reviews, page, PER_PAGE)
        return jsonify({
            "status": "success",
            "last_updated": data_manager.data.get("last_updated"),
            "summary": data_manager.get_summary(reviews),
            **paged,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/recent")
def recent_reviews():
    """Every review published within the last N days, newest first.
    Slim rows (no captions/transcripts) so long windows stay light; the modal
    fetches the full record via /api/review on click."""
    try:
        days = max(1, min(int(request.args.get("days", 30)), 3650))
        data_manager.reload_from_disk()
        cutoff = datetime.now() - timedelta(days=days)
        rows = []
        for r in data_manager.data.get("reviews", []):
            pa = r.get("published_at") or ""
            try:
                dt = datetime.fromisoformat(pa.replace("Z", "+00:00")).replace(tzinfo=None)
            except ValueError:
                continue
            if dt < cutoff:
                continue
            rows.append({
                "url": r.get("url"), "title": r.get("title"),
                "title_en": r.get("title_en"), "year": r.get("year"),
                "published_at": pa, "channel_title": r.get("channel_title"),
                "source_name": r.get("source_name"), "media_type": r.get("media_type"),
                "view_count": r.get("view_count"), "like_count": r.get("like_count"),
                "ife_system": r.get("ife_system"),
                "airlines_mentioned": (r.get("airlines_mentioned") or [])[:1],
                "transcript_available": r.get("transcript_available"),
                "transcript_excerpt": (r.get("transcript_excerpt") or "")[:160],
                "chapters": [{"ife": True}] if any(c.get("ife") for c in (r.get("chapters") or [])) else [],
            })
        rows.sort(key=lambda x: x["published_at"], reverse=True)
        return jsonify({"status": "success", "days": days, "total": len(rows), "rows": rows})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


@app.route("/api/review")
def get_single_review():
    """Return one full review object by its URL (for opening in the in-app modal)."""
    try:
        url = (request.args.get("url") or "").strip()
        if not url:
            return jsonify({"status": "error", "error": "url required"}), 400
        data_manager.reload_from_disk()
        for r in data_manager.data.get("reviews", []):
            if r.get("url") == url:
                return jsonify({"status": "success", "review": r})
        return jsonify({"status": "error", "error": "not found"}), 404
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


@app.route("/api/review-system", methods=["POST"])
def set_review_system():
    """Manually set or clear a review's IFE system tag.
    Only allowed from the host machine — teammates on the LAN get a 403."""
    if not _request_is_local():
        return jsonify({"status": "error",
                        "error": "editing is only allowed from the host machine"}), 403
    try:
        body = request.get_json() or {}
        url = (body.get("url") or "").strip()
        system = (body.get("system") or "").strip()
        if not url:
            return jsonify({"status": "error", "error": "url required"}), 400
        data_manager.reload_from_disk()
        for r in data_manager.data.get("reviews", []):
            if r.get("url") == url:
                r["ife_system"] = system or None
                r["ife_system_inferred"] = False
                r["ife_system_manual"] = True
                data_manager.save_cache()
                return jsonify({"status": "success", "ife_system": r["ife_system"]})
        return jsonify({"status": "error", "error": "not found"}), 404
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


@app.route("/api/ife-filter-options")
def get_filter_options():
    try:
        data_manager.reload_from_disk()
        return jsonify({
            "status": "success",
            "options": data_manager.get_filter_options()
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/ife-filter", methods=["POST"])
def filter_ife_reviews():
    """Filter + paginate reviews. Accepts optional _search for title text search."""
    try:
        body = request.get_json() or {}
        page   = int(body.pop("page", 1))
        search = body.pop("_search", "").strip().lower()
        channel = body.pop("_channel", "").strip().lower()

        data_manager.reload_from_disk()
        filtered = data_manager.filter_reviews(body)

        if channel:
            filtered = [r for r in filtered
                        if (r.get("channel_title") or "").strip().lower() == channel]

        if search:
            notes = _load_notes()
            def _matches(r):
                if search in (r.get("title") or "").lower():
                    return True
                if search in (r.get("transcript_excerpt") or "").lower():
                    return True
                if any(search in (c.get("text") or "").lower() for c in r.get("captions") or []):
                    return True
                if any(search in (c.get("title") or "").lower() for c in r.get("chapters") or []):
                    return True
                for n in notes.get(r.get("url", ""), []):
                    if search in (n.get("text", "") + " " + n.get("author", "")).lower():
                        return True
                return False
            filtered = [r for r in filtered if _matches(r)]

        paged = data_manager.paginate(filtered, page, PER_PAGE)
        return jsonify({
            "status": "success",
            "last_updated": data_manager.data.get("last_updated"),
            "summary": data_manager.get_summary(filtered),
            **paged,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/ife-seed", methods=["POST"])
def ife_seed():
    """Manually trigger a 100-result discovery crawl in the background."""
    def _run():
        try:
            data_manager.reload_from_disk()
            existing = {r["url"] for r in data_manager.data.get("reviews", [])}
            crawler  = IFECrawler(verify_ssl=False, api_key=os.environ.get("YOUTUBE_API_KEY", ""))
            results  = crawler.auto_discover(existing_urls=existing, max_results=500, days_lookback=365)
            if results:
                data_manager.reload_from_disk()
                data_manager.data["reviews"].extend(results)
                data_manager.save_cache()
        except Exception as e:
            print(f"[manual-seed] {e}")
    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "success", "message": "seed crawl started in background"})


@app.route("/api/ife-refresh")
def ife_refresh():
    """Lightweight poll: last_updated + total count + crawl status."""
    try:
        data_manager.reload_from_disk()
        return jsonify({
            "status":       "success",
            "last_updated": data_manager.data.get("last_updated"),
            "total":        len(data_manager.data.get("reviews", [])),
            "crawl":        _crawl_status,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/stats")
def stats():
    data_manager.reload_from_disk()
    reviews = data_manager.data.get("reviews", [])
    total = len(reviews)
    with_transcript = sum(1 for r in reviews if r.get("transcript_available"))
    by_source = {}
    by_system = {}
    by_airline = {}
    by_year = {}
    by_sentiment = {"positive": 0, "neutral": 0, "negative": 0}

    for r in reviews:
        src = r.get("source_name", "Creator")
        by_source[src] = by_source.get(src, 0) + 1
        sys = r.get("ife_system")
        if sys:
            by_system[sys] = by_system.get(sys, 0) + 1
        for a in r.get("airlines_mentioned", []):
            k = a["keyword"]
            by_airline[k] = by_airline.get(k, 0) + 1
        yr = r.get("year")
        if yr:
            by_year[str(yr)] = by_year.get(str(yr), 0) + 1
        sent = r.get("sentiment", "neutral")
        if sent in by_sentiment:
            by_sentiment[sent] += 1

    top_systems = sorted(by_system.items(), key=lambda x: x[1], reverse=True)[:12]
    top_airlines = sorted(by_airline.items(), key=lambda x: x[1], reverse=True)[:15]
    years_sorted = sorted(by_year.items())
    max_sys = top_systems[0][1] if top_systems else 1

    return render_template("stats.html",
        total=total,
        with_transcript=with_transcript,
        by_source=by_source,
        top_systems=top_systems,
        top_airlines=top_airlines,
        years=years_sorted,
        sentiment=by_sentiment,
        max_sys=max_sys,
        last_updated=data_manager.data.get("last_updated", ""),
    )


@app.route("/export.csv")
def export_csv():
    data_manager.reload_from_disk()
    # Optional query params mirror the Reviews tab (facets + search + channel),
    # so the download matches what's currently on screen.
    filters = {}
    for k in ("airlines", "aircraft", "ife_systems", "ife_features",
              "media_types", "source_tiers", "transcript", "chapters"):
        v = request.args.get(k)
        if v:
            filters[k] = v.split(",")
    if request.args.get("years"):
        filters["years"] = [int(y) for y in request.args["years"].split(",")
                            if y.strip().isdigit()]
    reviews = (data_manager.filter_reviews(filters) if filters
               else data_manager.data.get("reviews", []))

    channel = (request.args.get("channel") or "").strip().lower()
    if channel:
        reviews = [r for r in reviews
                   if (r.get("channel_title") or "").strip().lower() == channel]

    q = (request.args.get("q") or "").strip().lower()
    if q:
        notes = _load_notes()
        def _matches(r):
            if q in (r.get("title") or "").lower():
                return True
            if q in (r.get("transcript_excerpt") or "").lower():
                return True
            if any(q in (c.get("text") or "").lower() for c in r.get("captions") or []):
                return True
            if any(q in (c.get("title") or "").lower() for c in r.get("chapters") or []):
                return True
            for n in notes.get(r.get("url", ""), []):
                if q in (n.get("text", "") + " " + n.get("author", "")).lower():
                    return True
            return False
        reviews = [r for r in reviews if _matches(r)]

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["title", "url", "year", "source", "ife_system", "inferred",
                "airlines", "features", "views", "likes", "transcript_excerpt"])
    for r in reviews:
        w.writerow([
            r.get("title", ""),
            r.get("url", ""),
            r.get("year", ""),
            r.get("source_name", ""),
            r.get("ife_system", ""),
            "yes" if r.get("ife_system_inferred") else "no",
            "|".join(a["keyword"] for a in r.get("airlines_mentioned", [])),
            "|".join(r.get("ife_features", {}).keys()),
            r.get("view_count", ""),
            r.get("like_count", ""),
            (r.get("transcript_excerpt") or "")[:300],
        ])
    return Response(
        "﻿" + out.getvalue(),  # BOM for Excel UTF-8
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=ife_reviews.csv"},
    )


# ── Aggregates for Systems / Airlines / Analytics dashboards ───────────────────

SYSTEM_VENDORS = {
    "Emirates ICE": "Emirates / Thales", "Panasonic Astrova": "Panasonic Avionics",
    "Panasonic eX3": "Panasonic Avionics", "Panasonic eX2": "Panasonic Avionics",
    "Panasonic eX1": "Panasonic Avionics", "Thales AVANT Up": "Thales InFlyt",
    "Thales AVANT": "Thales InFlyt", "Safran RAVE Ultra": "Safran Passenger Innovations",
    "Safran RAVE": "Safran Passenger Innovations", "Collins Venue": "Collins Aerospace",
    "Viasat (streaming)": "Viasat", "Inmarsat GX": "Inmarsat", "Oryx One": "Qatar Airways",
    "KrisWorld": "Singapore Airlines", "StudioCX": "Cathay Pacific", "Lumexis FTTS": "Lumexis",
    "Gogo Avance": "Gogo", "Anuvu": "Anuvu", "Immfly": "Immfly", "Bluebox Wow": "Bluebox",
}

AIRLINE_REGIONS = {
    "emirates": "Middle East", "qatar airways": "Middle East", "etihad": "Middle East", "oman air": "Middle East",
    "ana": "Asia-Pacific", "japan airlines": "Asia-Pacific", "jal": "Asia-Pacific", "cathay pacific": "Asia-Pacific",
    "singapore airlines": "Asia-Pacific", "korean air": "Asia-Pacific", "eva air": "Asia-Pacific",
    "thai airways": "Asia-Pacific", "china airlines": "Asia-Pacific", "china eastern": "Asia-Pacific",
    "china southern": "Asia-Pacific", "hainan airlines": "Asia-Pacific", "air india": "Asia-Pacific",
    "qantas": "Asia-Pacific", "air new zealand": "Asia-Pacific",
    "lufthansa": "Europe", "british airways": "Europe", "air france": "Europe", "klm": "Europe",
    "turkish airlines": "Europe", "finnair": "Europe", "virgin atlantic": "Europe", "iberia": "Europe",
    "tap air portugal": "Europe", "swiss": "Europe", "austrian airlines": "Europe", "icelandair": "Europe",
    "wizz air": "Europe", "ryanair": "Europe", "easyjet": "Europe", "level": "Europe",
    "united airlines": "North America", "american airlines": "North America", "delta": "North America",
    "southwest": "North America", "alaska airlines": "North America", "air canada": "North America",
    "sun country": "North America", "frontier": "North America", "allegiant": "North America", "volaris": "North America",
}

_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

FEAT_LABELS = {
    "entertainment_system": "IFE System", "connectivity": "In-flight WiFi", "wifi": "In-flight WiFi",
    "4k_display": "4K HDR Display", "bluetooth_audio": "Bluetooth Audio", "content": "Content Library",
    "quality": "Display Quality", "seat": "Seat Comfort", "usb_power": "USB Power",
    "watch_party": "Watch Party", "seat_chat": "Seat-to-Seat Chat", "search": "Content Search",
    "tail_camera": "Tail / External Camera", "moving_map": "Moving Map",
}

# Tags that describe the flight/seat rather than an IFE capability — kept as
# review tags and filters, but excluded from "Popular Features" style lists.
NON_IFE_FEATURES = {"entertainment_system", "seat"}


def feat_label(k):
    return FEAT_LABELS.get(k, k.replace("_", " ").title())


def _aggregate_systems(reviews):
    agg = {}
    latest_year = max((r.get("year") for r in reviews if r.get("year")), default=None)
    for r in reviews:
        name = r.get("ife_system")
        if not name:
            continue
        a = agg.setdefault(name, {"reviews": 0, "views": 0, "likes": 0, "lc": 0,
                                  "features": set(), "airlines": {}, "aircraft": set(), "years": set()})
        a["reviews"] += 1
        a["views"] += int(r.get("view_count") or 0)
        if r.get("like_count") is not None:
            a["likes"] += r["like_count"]
            a["lc"] += 1
        a["features"].update(r.get("ife_features", {}).keys())
        for al in r.get("airlines_mentioned", []):
            a["airlines"][al["keyword"]] = a["airlines"].get(al["keyword"], 0) + 1
        a["aircraft"].update(ac["keyword"] for ac in r.get("aircraft_mentioned", []))
        if r.get("year"):
            a["years"].add(r["year"])
    out = []
    for name, a in agg.items():
        out.append({
            "name": name, "vendor": SYSTEM_VENDORS.get(name, "—"),
            "reviews": a["reviews"], "views": a["views"],
            "avg_likes": round(a["likes"] / a["lc"]) if a["lc"] else 0,
            "features": sorted(a["features"]),
            # most-mentioned airlines first (was alphabetical, which cut off frequent ones)
            "airlines": [k for k, _ in sorted(a["airlines"].items(), key=lambda x: -x[1])][:6],
            "aircraft": sorted(a["aircraft"])[:6],
            "status": "Active" if (latest_year and latest_year in a["years"]) else "Limited",
        })
    out.sort(key=lambda x: x["reviews"], reverse=True)
    return out


# Known airline↔system pairings the review corpus hasn't captured yet
# (e.g. Air France's A350 fleet runs Safran RAVE). Merged into the Airlines table.
AIRLINE_EXTRA_SYSTEMS = {
    "air france": ["Safran RAVE"],
}


def _aggregate_airlines(reviews):
    agg = {}
    for r in reviews:
        for al in r.get("airlines_mentioned", []):
            k = al["keyword"]
            a = agg.setdefault(k, {"reviews": 0, "views": 0, "systems": {}})
            a["reviews"] += 1
            a["views"] += int(r.get("view_count") or 0)
            sysname = r.get("ife_system")
            if sysname:
                a["systems"][sysname] = a["systems"].get(sysname, 0) + 1
    out = []
    for k, a in agg.items():
        # Airlines often run more than one system across fleets — show the top
        # ones seen in reviews, plus the known lookup if reviews missed it.
        ranked = [s for s, _ in sorted(a["systems"].items(), key=lambda x: -x[1])]
        lookup = AIRLINE_IFE_LOOKUP.get(k)
        if lookup and lookup not in ranked:
            ranked.append(lookup)
        for extra in AIRLINE_EXTRA_SYSTEMS.get(k, []):
            if extra not in ranked:
                ranked.insert(0, extra)
        out.append({
            "name": k, "reviews": a["reviews"],
            "system": ", ".join(ranked[:3]) if ranked else "—",
            "region": AIRLINE_REGIONS.get(k, "—"), "views": a["views"],
        })
    out.sort(key=lambda x: x["reviews"], reverse=True)
    return out[:30]


def _analytics(reviews, systems):
    # Per-feature review counts (how often each feature is discussed)
    feat = {}
    for r in reviews:
        for f in r.get("ife_features", {}).keys():
            if f not in NON_IFE_FEATURES:
                feat[f] = feat.get(f, 0) + 1
    features = sorted(({"key": k, "reviews": c} for k, c in feat.items()),
                      key=lambda x: x["reviews"], reverse=True)
    # Feature coverage: how many tracked systems expose each feature
    total_sys = len(systems) or 1
    fcov = {}
    for s in systems:
        for f in s["features"]:
            fcov[f] = fcov.get(f, 0) + 1
    feature_coverage = sorted(
        ({"key": k, "count": c, "total": total_sys} for k, c in fcov.items()),
        key=lambda x: x["count"], reverse=True,
    )[:6]
    # Review volume by year
    by_year = {}
    for r in reviews:
        y = r.get("year")
        if y:
            by_year[str(y)] = by_year.get(str(y), 0) + 1
    return {
        "by_system": [[s["name"], s["reviews"]] for s in systems[:8]],
        "by_year": sorted(by_year.items()),
        "features": features[:6],
        "feature_coverage": feature_coverage,
    }


def _regional(reviews):
    reg = {}
    for r in reviews:
        als = r.get("airlines_mentioned") or []
        if not als:
            continue
        region = AIRLINE_REGIONS.get(als[0].get("keyword"))
        if not region:
            continue
        d = reg.setdefault(region, {"reviews": 0, "views": 0})
        d["reviews"] += 1
        d["views"] += int(r.get("view_count") or 0)
    return sorted(({"region": k, **v} for k, v in reg.items()),
                  key=lambda x: x["reviews"], reverse=True)


def _momentum(reviews):
    """Per-system review growth, latest year vs prior — 'what's rising'."""
    sysyear, years = {}, set()
    for r in reviews:
        s, y = r.get("ife_system"), r.get("year")
        if not s or not y:
            continue
        sysyear.setdefault(s, {})
        sysyear[s][y] = sysyear[s].get(y, 0) + 1
        years.add(y)
    if not years:
        return {"year": None, "prev": None, "systems": []}
    cur = max(years)
    prev = cur - 1
    out = []
    for s, ys in sysyear.items():
        c, p = ys.get(cur, 0), ys.get(prev, 0)
        out.append({"name": s, "cur": c, "prev": p, "delta": c - p,
                    "pct": round((c - p) / p * 100) if p else (100 if c else 0)})
    out.sort(key=lambda x: (x["delta"], x["cur"]), reverse=True)
    return {"year": cur, "prev": prev, "systems": out}


def _coverage_gaps(systems):
    fcount = {}
    for s in systems:
        for f in s["features"]:
            fcount[f] = fcount.get(f, 0) + 1
    common = [f for f, c in fcount.items() if c >= 3]
    gaps = []
    for s in systems:
        missing = [f for f in common if f not in s["features"]]
        if missing:
            gaps.append({"system": s["name"], "missing": missing})
    stale = [{"name": s["name"], "reviews": s["reviews"]} for s in systems if s["status"] != "Active"]
    return {"stale": stale, "feature_gaps": gaps[:12], "common_features": common}


def _feature_mentions(reviews):
    """How often each feature is actually *discussed* in transcripts (not just tagged)."""
    out = {}
    for r in reviews:
        parts = [r.get("transcript_excerpt") or ""]
        parts += [c.get("text", "") for c in (r.get("captions") or [])]
        text = " ".join(parts).lower()
        if not text.strip():
            continue
        for key, kws in IFE_FEATURE_KEYWORDS.items():
            if key in NON_IFE_FEATURES:
                continue
            hits = sum(text.count(kw) for kw in kws)
            if hits:
                d = out.setdefault(key, {"reviews": 0, "mentions": 0})
                d["reviews"] += 1
                d["mentions"] += hits
    return sorted(({"key": k, **v} for k, v in out.items()),
                  key=lambda x: x["mentions"], reverse=True)


def _channel_list(reviews):
    stats = {}
    p = Path(__file__).parent / "channel_stats.json"
    if p.exists():
        stats = json.loads(p.read_text(encoding="utf-8"))
    title2key = {}
    for k, s in stats.items():
        title2key[s.get("title", "").lower()] = k
        title2key[k.lower()] = k
    agg = {}
    for r in reviews:
        k = title2key.get((r.get("channel_title") or "").lower())
        if not k:
            continue
        a = agg.setdefault(k, {"n": 0, "likes": 0, "lc": 0, "views": 0})
        a["n"] += 1
        a["views"] += int(r.get("view_count") or 0)
        if r.get("like_count") is not None:
            a["likes"] += r["like_count"]
            a["lc"] += 1
    out = []
    for k, s in stats.items():
        a = agg.get(k, {})
        out.append({
            "name": k, "title": s.get("title", k),
            "subscribers": s.get("subscribers", 0), "video_count": s.get("video_count", 0),
            "thumb": s.get("thumb", ""), "reviews": a.get("n", 0), "views": a.get("views", 0),
            "avg_likes": round(a["likes"] / a["lc"]) if a.get("lc") else None,
        })
    # Only channels that actually have IFE reviews in the library — a channel
    # with zero reviews has nothing to click through to.
    out = [c for c in out if c["reviews"] > 0]
    out.sort(key=lambda x: x["subscribers"], reverse=True)
    return out


def _metrics(reviews):
    """Headline numbers + real month-over-month change from publish dates."""
    months = {}
    for r in reviews:
        pa = r.get("published_at") or ""
        if len(pa) >= 7 and pa[4] == "-":
            months[pa[:7]] = months.get(pa[:7], 0) + 1
    keys = sorted(months)
    this_m = months[keys[-1]] if keys else 0
    last_m = months[keys[-2]] if len(keys) >= 2 else 0
    mom = round((this_m - last_m) / last_m * 100) if last_m else 0
    return {
        "total": len(reviews),
        "with_transcript": sum(1 for r in reviews if r.get("transcript_available")),
        "new_this_month": this_m,
        "last_month": last_m,
        "mom_pct": mom,
        "has_mom": last_m > 0,
    }


@app.route("/api/ife-aggregates")
def ife_aggregates():
    try:
        data_manager.reload_from_disk()
        reviews = data_manager.data.get("reviews", [])
        systems = _aggregate_systems(reviews)
        airlines = _aggregate_airlines(reviews)
        return jsonify({
            "status": "success",
            "metrics": _metrics(reviews),
            "systems": systems,
            "airlines": airlines,
            "analytics": _analytics(reviews, systems),
            "regional": _regional(reviews),
            "momentum": _momentum(reviews),
            "coverage_gaps": _coverage_gaps(systems),
            "feature_mentions": _feature_mentions(reviews)[:8],
            "reach": {"total_views": sum(int(r.get("view_count") or 0) for r in reviews)},
        })
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


@app.route("/api/ife-feature")
def ife_feature_detail():
    """Breakdown for a single IFE feature: headline stats, which systems and
    airlines it shows up with, review volume by year, and the feature's videos
    sorted by most popular (views)."""
    try:
        key = (request.args.get("key") or "").strip()
        if not key:
            return jsonify({"status": "error", "error": "key required"}), 400
        data_manager.reload_from_disk()
        reviews = data_manager.data.get("reviews", [])
        tagged = [r for r in reviews if key in (r.get("ife_features") or {})]
        if not tagged:
            return jsonify({"status": "error", "error": "no reviews for this feature"}), 404

        systems, airlines, by_year = {}, {}, {}
        likes, lc = 0, 0
        for r in tagged:
            s = r.get("ife_system")
            if s:
                systems[s] = systems.get(s, 0) + 1
            for al in r.get("airlines_mentioned", []):
                airlines[al["keyword"]] = airlines.get(al["keyword"], 0) + 1
            if r.get("year"):
                by_year[str(r["year"])] = by_year.get(str(r["year"]), 0) + 1
            if r.get("like_count") is not None:
                likes += r["like_count"]
                lc += 1

        # How often the feature is actually discussed (same method as _feature_mentions)
        kws = IFE_FEATURE_KEYWORDS.get(key, [])
        mentions = 0
        for r in tagged:
            parts = [r.get("transcript_excerpt") or ""]
            parts += [c.get("text", "") for c in (r.get("captions") or [])]
            text = " ".join(parts).lower()
            mentions += sum(text.count(kw) for kw in kws)

        videos = sorted(tagged, key=lambda r: int(r.get("view_count") or 0), reverse=True)
        return jsonify({
            "status": "success",
            "key": key,
            "label": feat_label(key),
            "stats": {
                "reviews": len(tagged),
                "total_views": sum(int(r.get("view_count") or 0) for r in tagged),
                "avg_likes": round(likes / lc) if lc else 0,
                "mentions": mentions,
            },
            "systems": sorted(({"name": k, "reviews": c} for k, c in systems.items()),
                              key=lambda x: x["reviews"], reverse=True),
            "airlines": sorted(({"name": k, "reviews": c} for k, c in airlines.items()),
                               key=lambda x: x["reviews"], reverse=True)[:8],
            "by_year": sorted(by_year.items()),
            "videos": videos[:40],
        })
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


@app.route("/api/channels")
def channels():
    """Known independent reviewer channels: subscribers (from channel_stats.json)
    + avg likes / review count computed from the cached IFE reviews."""
    try:
        data_manager.reload_from_disk()
        return jsonify({"status": "success",
                        "channels": _channel_list(data_manager.data.get("reviews", []))})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


# Strong IFE signals only — a quote must clearly reference an IFE system, screen,
# feature, or hardware. Deliberately EXCLUDES loose words (movie, content, music,
# channel, audio, life…) that pulled in drama/vlog noise.
_IFE_COMMENT_TERMS = (
    "inflight entertainment", "in-flight entertainment", "in flight entertainment",
    "entertainment system", "entertainment screen", "entertainment offering",
    "entertainment control", "entertainment selection", "onboard entertainment",
    "seatback", "seat back", "seat-back", "seat screen", "seatback screen",
    "touchscreen", "touch screen", "touch-screen",
    "seatback pocket", "screen features", "inch screen", "inch touchscreen",
    "amoled", "oled", "4k", "hd screen", "uhd",
    "bluetooth", "headphone", "noise cancelling", "noise-cancelling",
    "wi-fi", "wifi", "starlink", "viasat", "onboard wifi",
    "usb", "charging port", "charging point", "power outlet", "power port",
    "panasonic", "thales", "safran", "rave", "astrova", "avant", "krisworld",
    "oryx", "emirates ice", "studiocx", "gogo", "entertainment",
)
# Match terms only at a word start (leading boundary) so "usb" doesn't match
# "hUSBand" and "rave" doesn't match "tRAVEl"; trailing is free so plurals
# ("headphones", "screens") still match.
_IFE_TERM_RE = re.compile(
    r"(?<![a-z0-9])(?:" + "|".join(re.escape(t) for t in _IFE_COMMENT_TERMS + ("ife",)) + r")",
    re.I,
)


# Phrases that contain an IFE term but describe seat furniture, not the IFE
# (e.g. "to my left, I've got a headphone hook").
_IFE_FALSE_POSITIVES = ("headphone hook", "headphone hanger", "headphone stand", "coat hook")


def _is_ife_quote(text):
    low = (text or "").lower()
    for fp in _IFE_FALSE_POSITIVES:
        low = low.replace(fp, "")
    return bool(_IFE_TERM_RE.search(low))


@app.route("/api/comments")
def ife_comments():
    """Real IFE-related YouTube comments (channel + viewers), harvested per video
    via the Data API (see gather_comments.py). Falls back to IFE transcript quotes
    for videos whose comments haven't been gathered yet."""
    try:
        data_manager.reload_from_disk()
        reviews = sorted(data_manager.data.get("reviews", []),
                         key=lambda r: (r.get("year") or 0, r.get("view_count") or 0), reverse=True)
        out, seen = [], set()
        for r in reviews:
            channel = r.get("channel_title") or r.get("source_name") or "Reviewer"
            on = ((r.get("airlines_mentioned") or [{}])[0].get("keyword")
                  or r.get("ife_system") or "IFE")
            vbase = {"on": on, "url": r.get("url", ""),
                     "year": r.get("year"), "system": r.get("ife_system"),
                     "channel": channel}

            yc = r.get("yt_comments") or []
            if yc:
                # Real YouTube comments: the uploader's own comments and viewers'.
                added = 0
                for c in yc:
                    t = (c.get("text") or "").strip()
                    if len(t) < 20:
                        continue
                    key = t.lower()[:70]
                    if key in seen:
                        continue
                    seen.add(key)
                    out.append({**vbase, "who": c.get("author") or "Viewer",
                                "text": t, "likes": c.get("likes") or 0,
                                "when": c.get("when"),
                                "text_en": c.get("text_en"), "lang": c.get("lang"),
                                "is_channel": bool(c.get("is_channel")),
                                "kind": "comment",
                                "timestamp": None, "start": None})
                    added += 1
                    if added >= 3:  # keep the feed diverse across videos
                        break
                if added:
                    if len(out) >= 200:
                        break
                    continue

            # Fallback: IFE-relevant line straight from the transcript.
            ex = (r.get("transcript_excerpt") or "").strip()
            cap0 = next((c.get("text", "").strip() for c in (r.get("captions") or [])
                         if len((c.get("text") or "").strip()) >= 25
                         and _is_ife_quote(c.get("text", ""))), "")
            quote = cap0 or (ex if len(ex) >= 25 and _is_ife_quote(ex) else "")
            if quote:
                key = quote.lower()[:70]
                if key not in seen:
                    seen.add(key)
                    out.append({**vbase, "who": channel, "text": quote,
                                "likes": 0, "when": None, "is_channel": True,
                                "kind": "transcript", "timestamp": None, "start": None})
            if len(out) >= 200:
                break
        return jsonify({"status": "success", "comments": out, "count": len(out)})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


@app.route("/api/crawl")
def crawl_detail():
    """Crawl monitor: live status + parsed tail of the crawl log."""
    try:
        data_manager.reload_from_disk()
        reviews = data_manager.data.get("reviews", [])
        pending = sum(1 for r in reviews
                      if r.get("media_type") == "video" and not r.get("transcript_available"))
        log = []
        p = Path(__file__).parent / "crawl_log.txt"
        if p.exists():
            lines = p.read_text(encoding="utf-8", errors="ignore").splitlines()
            for ln in reversed(lines[-60:]):
                ln = ln.strip()
                if not ln:
                    continue
                parts = ln.split(None, 3)
                tm, lvl, msg = "", "INFO", ln
                if len(parts) >= 3 and ":" in parts[1]:
                    tm = parts[1][:5]  # HH:MM from the time token
                    lvl = parts[2]
                    msg = parts[3] if len(parts) > 3 else ""
                log.append({
                    "time": tm,
                    "status": "warn" if lvl in ("WARNING", "ERROR") else "ok",
                    "msg": msg or ln,
                })
                if len(log) >= 14:
                    break
        return jsonify({
            "status": "success",
            "crawl": _crawl_status,
            "total": len(reviews),
            "pending_transcripts": pending,
            "queries_total": len(getattr(IFECrawler, "YOUTUBE_QUERIES", []) or []) or 65,
            "log": log,
        })
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


# ── AI chat (natural-language search + insight over the review corpus) ─────────

CHAT_MODEL = os.environ.get("CHAT_MODEL", "claude-fable-5")
_STOPWORDS = {
    "the", "a", "an", "and", "or", "of", "to", "in", "on", "for", "is", "are",
    "what", "which", "how", "do", "does", "tell", "me", "show", "about", "any",
    "with", "that", "this", "have", "has", "i", "want", "videos", "video",
    "review", "reviews", "give", "general", "if", "up", "whats", "its",
}


def _tokens(text: str):
    return {w for w in re.findall(r"[a-z0-9]+", (text or "").lower()) if len(w) > 2 and w not in _STOPWORDS}


def _search_text(r: dict) -> str:
    parts = [
        r.get("title", ""),
        r.get("ife_system", "") or "",
        r.get("channel_title", "") or "",
        r.get("transcript_excerpt", "") or "",
        " ".join(a.get("keyword", "") for a in r.get("airlines_mentioned", [])),
        " ".join(a.get("keyword", "") for a in r.get("aircraft_mentioned", [])),
        " ".join(r.get("ife_features", {}).keys()),
        " ".join(c.get("text", "") for c in r.get("captions", []) or []),
    ]
    return " ".join(parts)


def _rank_reviews(question: str, reviews: list, limit: int = 40) -> list:
    """Lexical pre-filter so we only send the most relevant reviews to the model."""
    q = _tokens(question)
    if not q:
        return sorted(reviews, key=data_manager._relevance, reverse=True)[:limit]
    scored = []
    for r in reviews:
        overlap = len(q & _tokens(_search_text(r)))
        if overlap:
            scored.append((overlap + data_manager._relevance(r), r))
    scored.sort(key=lambda x: x[0], reverse=True)
    ranked = [r for _, r in scored[:limit]]
    if len(ranked) < 8:  # weak lexical match — top up with most relevant overall
        seen = {id(r) for r in ranked}
        for r in sorted(reviews, key=data_manager._relevance, reverse=True):
            if id(r) not in seen:
                ranked.append(r)
            if len(ranked) >= 8:
                break
    return ranked


def _digest(reviews: list) -> str:
    """Compact one-line-per-review context block the model reasons over."""
    lines = []
    for i, r in enumerate(reviews):
        airlines = ", ".join(a.get("keyword", "") for a in r.get("airlines_mentioned", [])[:3])
        feats = ", ".join(r.get("ife_features", {}).keys())
        excerpt = (r.get("transcript_excerpt") or "")[:220]
        lines.append(
            f"[{i}] {r.get('title', 'Untitled')}\n"
            f"    system={r.get('ife_system') or 'unknown'} | airlines={airlines or '-'} | "
            f"year={r.get('year') or '-'} | sentiment={r.get('sentiment', 'neutral')} | "
            f"features={feats or '-'}\n"
            f"    excerpt: {excerpt}"
        )
    return "\n".join(lines)


CHAT_SYSTEM = (
    "You are the analyst for IFE ReviewDB, a database of in-flight entertainment (IFE) "
    "reviews (mostly YouTube videos and some articles). Answer the user's question using "
    "ONLY the reviews provided in the context block. Each review has a numbered index [n].\n\n"
    "Lead with the answer. Synthesize across reviews — what a manufacturer (Panasonic, Thales, "
    "Safran) is doing, which airlines/aircraft run a system, and which features (4K, Bluetooth, "
    "wifi) show up. This is a statistics and information tool: report facts, counts, and coverage — "
    "do NOT rate or score sentiment. Be concrete and concise; use short markdown. "
    "If the reviews don't cover the question, say so plainly rather than inventing facts. "
    "Cite the specific reviews you drew on by their index in source_indexes."
)

CHAT_SCHEMA = {
    "type": "object",
    "properties": {
        "answer": {"type": "string", "description": "Markdown answer, lead with the conclusion."},
        "source_indexes": {
            "type": "array",
            "items": {"type": "integer"},
            "description": "Indexes [n] of the reviews actually used in the answer.",
        },
    },
    "required": ["answer", "source_indexes"],
    "additionalProperties": False,
}


@app.route("/api/chat", methods=["POST"])
def chat():
    body = request.get_json() or {}
    question = (body.get("question") or "").strip()
    if not question:
        return jsonify({"status": "error", "error": "Ask a question first."}), 400
    if not _ANTHROPIC_OK:
        return jsonify({"status": "error", "error": "AI chat is unavailable — the 'anthropic' package is not installed."}), 503

    try:
        client = anthropic.Anthropic()  # resolves ANTHROPIC_API_KEY / profile from env
    except Exception:
        return jsonify({"status": "error", "error": "AI chat is not configured — set ANTHROPIC_API_KEY."}), 503

    data_manager.reload_from_disk()
    reviews = data_manager.data.get("reviews", [])
    ranked = _rank_reviews(question, reviews)

    req = dict(
        model=CHAT_MODEL,
        max_tokens=1500,
        system=CHAT_SYSTEM,
        thinking={"type": "adaptive"},
        output_config={"effort": "medium", "format": {"type": "json_schema", "schema": CHAT_SCHEMA}},
        messages=[{"role": "user", "content": f"Question: {question}\n\nReviews:\n{_digest(ranked)}"}],
    )
    is_fable = CHAT_MODEL.startswith(("claude-fable", "claude-mythos"))
    try:
        if is_fable:
            # Fable/Mythos: safety classifiers may decline — auto-fall back to Opus 4.8.
            resp = client.beta.messages.create(
                betas=["server-side-fallback-2026-06-01"],
                fallbacks=[{"model": "claude-opus-4-8"}],
                **req,
            )
        else:
            resp = client.messages.create(**req)
    except anthropic.AuthenticationError:
        return jsonify({"status": "error", "error": "AI chat auth failed — check ANTHROPIC_API_KEY."}), 503
    except Exception as e:
        msg = str(e).lower()
        if "authentication method" in msg or "api_key" in msg or "x-api-key" in msg:
            return jsonify({"status": "error", "error": "AI chat is not configured — set ANTHROPIC_API_KEY in your .env."}), 503
        if "retention" in msg:
            return jsonify({"status": "error", "error": "Claude Fable 5 requires 30-day data retention — enable it, or set CHAT_MODEL=claude-opus-4-8."}), 502
        return jsonify({"status": "error", "error": f"AI request failed: {e}"}), 502

    if getattr(resp, "stop_reason", None) == "refusal":
        return jsonify({"status": "success", "considered": len(ranked), "sources": [],
                        "answer": "The model declined to answer that request."})

    text = next((b.text for b in resp.content if b.type == "text"), "")
    try:
        parsed = json.loads(text)
    except (ValueError, TypeError):
        parsed = {"answer": text, "source_indexes": []}

    sources = []
    for i in parsed.get("source_indexes", []):
        if isinstance(i, int) and 0 <= i < len(ranked):
            r = ranked[i]
            sources.append({
                "title": r.get("title", "Untitled"),
                "url": r.get("url", ""),
                "ife_system": r.get("ife_system"),
                "year": r.get("year"),
                "media_type": r.get("media_type", "video"),
            })

    return jsonify({
        "status": "success",
        "answer": parsed.get("answer", ""),
        "sources": sources,
        "considered": len(ranked),
    })


# ── Shared notes (persisted to notes.json, keyed by review URL) ────────────────

NOTES_FILE = Path(__file__).parent / "notes.json"
_notes_lock = threading.Lock()


def _load_notes():
    if NOTES_FILE.exists():
        try:
            return json.loads(NOTES_FILE.read_text(encoding="utf-8"))
        except (ValueError, OSError):
            return {}
    return {}


def _save_notes(notes):
    with _notes_lock:
        NOTES_FILE.write_text(json.dumps(notes, indent=2, ensure_ascii=False), encoding="utf-8")


@app.route("/api/notes", methods=["GET"])
def get_notes():
    notes = _load_notes()
    url = request.args.get("url")
    if url:
        return jsonify({"status": "success", "notes": notes.get(url, [])})
    return jsonify({"status": "success", "notes": notes})


@app.route("/api/notes", methods=["POST"])
def add_note():
    body = request.get_json() or {}
    url = (body.get("url") or "").strip()
    text = (body.get("text") or "").strip()
    author = (body.get("author") or "Anonymous").strip()[:80] or "Anonymous"
    if not url or not text:
        return jsonify({"status": "error", "error": "A note and its video are required."}), 400
    with _notes_lock:
        notes = _load_notes()
        entry = {"author": author, "text": text[:2000], "ts": datetime.now().isoformat()}
        notes.setdefault(url, []).append(entry)
        NOTES_FILE.write_text(json.dumps(notes, indent=2, ensure_ascii=False), encoding="utf-8")
    return jsonify({"status": "success", "notes": notes[url]})


# ── Internal reviews: manual entry + MS Forms Excel import ─────────────────────

def _parse_review_date(v):
    """Best-effort date → (year, iso_string). Accepts datetime, ISO, US/EU forms."""
    if isinstance(v, datetime):
        return v.year, v.strftime("%Y-%m-%dT%H:%M:%SZ")
    s = str(v or "").strip()
    if not s:
        return None, None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m/%d/%y", "%Y/%m/%d", "%B %Y", "%b %Y", "%Y"):
        try:
            dt = datetime.strptime(s[:19].split("T")[0] if "T" in s else s, fmt)
            return dt.year, dt.strftime("%Y-%m-%dT00:00:00Z")
        except ValueError:
            continue
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.year, dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    except ValueError:
        return None, None


def _make_internal_review(title="", text="", airline="", aircraft="", system="",
                          rating=None, author="", date_val=None):
    """Build a review record from team input so it flows through the same
    dashboard as crawled reviews. URL is a stable hash so re-imports dedupe."""
    import hashlib
    text = (text or "").strip()
    author = (author or "").strip()
    airline = (airline or "").strip()
    aircraft = (aircraft or "").strip()
    title = (title or "").strip() or ("Internal review — " + (airline or "IFE") + (" " + aircraft if aircraft else ""))
    uid = hashlib.sha1((author + "|" + title + "|" + text[:100]).encode("utf-8", "ignore")).hexdigest()[:16]

    low = (title + " " + text).lower()
    feats = {f: True for f, kws in IFE_FEATURE_KEYWORDS.items() if any(kw in low for kw in kws)}
    airlines = ([{"keyword": airline.lower(), "mentions": 1}] if airline else [])
    for kw in AIRLINE_KEYWORDS:
        if kw in low and all(a["keyword"] != kw for a in airlines):
            airlines.append({"keyword": kw, "mentions": low.count(kw)})
    aircraft_m = ([{"keyword": aircraft.lower(), "mentions": 1}] if aircraft else [])
    for kw in AIRCRAFT_KEYWORDS:
        if kw in low and all(a["keyword"] != kw for a in aircraft_m):
            aircraft_m.append({"keyword": kw, "mentions": low.count(kw)})

    try:
        rating = int(rating) if rating not in (None, "") else None
        rating = max(1, min(5, rating)) if rating else None
    except (ValueError, TypeError):
        rating = None
    sentiment = "neutral"
    if rating:
        sentiment = "positive" if rating >= 4 else ("negative" if rating <= 2 else "neutral")

    year, published = _parse_review_date(date_val)
    if not year:
        now = datetime.now()
        year, published = now.year, now.strftime("%Y-%m-%dT%H:%M:%SZ")

    return {
        "url": f"internal://{uid}",
        "title": title[:150],
        "year": year,
        "published_at": published,
        "channel_title": author or "Team member",
        "view_count": 0, "like_count": 0,
        "ife_system": (system or "").strip() or None,
        "ife_system_manual": bool((system or "").strip()),
        "ife_system_inferred": False,
        "media_type": "internal",
        "airlines_mentioned": airlines[:5],
        "aircraft_mentioned": aircraft_m[:5],
        "ife_features": feats,
        "ife_specs": {},
        "transcript_available": False,
        "transcript_excerpt": text[:300],
        "internal_text": text,
        "internal_rating": rating,
        "internal_author": author,
        "captions": [], "chapters": [],
        "source_tier": 1, "source_name": "Internal",
        "sentiment": sentiment,
    }


@app.route("/api/internal-review", methods=["POST"])
def add_internal_review():
    """Manual entry of a team member's own IFE review."""
    try:
        b = request.get_json() or {}
        text = (b.get("text") or "").strip()
        if len(text) < 10:
            return jsonify({"status": "error", "error": "Please write at least a sentence about the IFE."}), 400
        rec = _make_internal_review(b.get("title"), text, b.get("airline"), b.get("aircraft"),
                                    b.get("system"), b.get("rating"), b.get("author"), b.get("date"))
        data_manager.reload_from_disk()
        if any(r.get("url") == rec["url"] for r in data_manager.data.get("reviews", [])):
            return jsonify({"status": "error", "error": "This review was already added."}), 409
        data_manager.data["reviews"].append(rec)
        data_manager.save_cache()
        return jsonify({"status": "success", "review": rec})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


def _map_forms_headers(headers):
    """Heuristically map MS Forms export columns to review fields."""
    m = {}
    for i, h in enumerate(headers):
        hl = str(h or "").strip().lower()
        if not hl:
            continue
        def put(key):
            m.setdefault(key, i)
        if "airline" in hl:
            put("airline")
        elif "system" in hl:
            # before the aircraft rule: "which IFE system did the aircraft have"
            put("system")
        elif "aircraft" in hl or "plane" in hl:
            put("aircraft")
        elif "rating" in hl or "rate" in hl or "score" in hl or "stars" in hl:
            put("rating")
        elif "email" in hl:
            put("email")
        elif hl == "name" or "your name" in hl:
            put("author")
        elif "completion time" in hl:
            put("completion")
        elif "date" in hl and "start" not in hl:
            put("date")
        elif "title" in hl:
            put("title")
        elif any(k in hl for k in ("review", "comment", "feedback", "experience",
                                   "thoughts", "describe", "opinion", "notes")):
            put("text")
    return m


@app.route("/api/import-forms", methods=["POST"])
def import_forms():
    """Import an MS Forms .xlsx export. Without commit=1 returns a mapped
    preview; with commit=1 appends the new (deduped) reviews."""
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"status": "error", "error": "no file uploaded"}), 400
        from openpyxl import load_workbook
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = [r for r in ws.iter_rows(values_only=True)]
        if len(rows) < 2:
            return jsonify({"status": "error", "error": "the sheet has no data rows"}), 400
        headers = list(rows[0])
        m = _map_forms_headers(headers)
        if "text" not in m:
            # fall back to the unmapped column with the longest average text
            used = set(m.values())
            best, best_len = None, 0
            for i in range(len(headers)):
                if i in used:
                    continue
                vals = [str(r[i]) for r in rows[1:] if i < len(r) and r[i]]
                avg = sum(len(v) for v in vals) / len(vals) if vals else 0
                if avg > best_len:
                    best, best_len = i, avg
            if best is None or best_len < 20:
                return jsonify({"status": "error", "error": "could not find a review-text column"}), 400
            m["text"] = best

        def cell(row, key):
            i = m.get(key)
            return row[i] if i is not None and i < len(row) else None

        recs = []
        for row in rows[1:]:
            text = str(cell(row, "text") or "").strip()
            if len(text) < 10:
                continue
            recs.append(_make_internal_review(
                title=str(cell(row, "title") or ""),
                text=text,
                airline=str(cell(row, "airline") or ""),
                aircraft=str(cell(row, "aircraft") or ""),
                system=str(cell(row, "system") or ""),
                rating=cell(row, "rating"),
                author=str(cell(row, "author") or "").strip() or str(cell(row, "email") or "").split("@")[0],
                date_val=cell(row, "date") or cell(row, "completion"),
            ))

        mapping_names = {k: str(headers[i]) for k, i in m.items()}
        if request.args.get("commit") != "1":
            prev = [{"title": r["title"], "airline": (r["airlines_mentioned"][:1] or [{}])[0].get("keyword", ""),
                     "system": r["ife_system"] or "", "author": r["internal_author"],
                     "year": r["year"], "rating": r["internal_rating"],
                     "text": r["internal_text"][:120]} for r in recs[:8]]
            return jsonify({"status": "success", "preview": prev, "total": len(recs), "mapping": mapping_names})

        data_manager.reload_from_disk()
        have = {r.get("url") for r in data_manager.data.get("reviews", [])}
        new = [r for r in recs if r["url"] not in have]
        data_manager.data["reviews"].extend(new)
        if new:
            data_manager.save_cache()
        return jsonify({"status": "success", "added": len(new), "skipped": len(recs) - len(new), "mapping": mapping_names})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


# ── Saved videos (persisted to flags.json, keyed by review URL) ────────────────
# Server-side so bookmarks survive browser changes and work from any device
# on the network (they used to live in per-browser localStorage).

FLAGS_FILE = Path(__file__).parent / "flags.json"
_flags_lock = threading.Lock()


def _load_flags():
    if FLAGS_FILE.exists():
        try:
            return json.loads(FLAGS_FILE.read_text(encoding="utf-8"))
        except (ValueError, OSError):
            return {}
    return {}


@app.route("/api/flags", methods=["GET"])
def get_flags():
    return jsonify({"status": "success", "flags": _load_flags()})


@app.route("/api/flags", methods=["POST"])
def set_flag():
    body = request.get_json() or {}
    url = (body.get("url") or "").strip()
    if not url:
        return jsonify({"status": "error", "error": "url required"}), 400
    with _flags_lock:
        flags = _load_flags()
        if body.get("on"):
            flags[url] = {
                "url": url,
                "title": (body.get("title") or "Untitled")[:200],
                "channel": (body.get("channel") or "")[:100],
                "year": body.get("year") or "",
                "ts": body.get("ts") or int(time.time() * 1000),
            }
        else:
            flags.pop(url, None)
        FLAGS_FILE.write_text(json.dumps(flags, indent=2, ensure_ascii=False), encoding="utf-8")
    return jsonify({"status": "success", "count": len(flags)})


@app.route("/report")
def report():
    """On-demand printable intelligence digest (print to PDF from the browser)."""
    data_manager.reload_from_disk()
    reviews = data_manager.data.get("reviews", [])
    systems = _aggregate_systems(reviews)
    return render_template(
        "report.html",
        generated=datetime.now().strftime("%B %d, %Y · %H:%M"),
        total=len(reviews),
        with_transcript=sum(1 for r in reviews if r.get("transcript_available")),
        total_views=sum(int(r.get("view_count") or 0) for r in reviews),
        systems=systems[:10],
        regional=_regional(reviews),
        momentum=_momentum(reviews),
        channels=[c for c in _channel_list(reviews) if c["reviews"] > 0][:10],
        gaps=_coverage_gaps(systems),
        feature_mentions=_feature_mentions(reviews)[:8],
        feat_label=feat_label,
    )


if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
